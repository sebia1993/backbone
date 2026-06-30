from __future__ import annotations

import csv
import json
from dataclasses import asdict
from html import escape
from pathlib import Path

from .models import DiffItem, DiffLine, DiffSummary
from .redaction import redact_payload, redact_sensitive_text
from .report_bundle import create_share_report_bundle
from .snapshot import SnapshotStore, sanitize_filename
from .version import APP_NAME, APP_VERSION
from .workflow import is_sample_snapshot, severity_to_korean, status_to_korean


SUMMARY_LABELS_KO = {
    "No meaningful change detected.": "의미 있는 변경 없음",
    "Target snapshot command failed.": "비교 스냅샷에서 명령 실행 실패",
    "Critical state keyword detected in changed output.": "변경 출력에서 긴급 상태 키워드 감지",
    "New critical-looking log line detected.": "신규 긴급 의심 로그 감지",
    "Warning keyword detected in changed output.": "변경 출력에서 주의 키워드 감지",
    "Operational state changed.": "운영 상태 변경 감지",
    "Output changed.": "출력 변경 감지",
    "Target device connection failed.": "비교 시점 장비 접속 실패",
    "Target device connection restored.": "비교 시점 장비 접속 복구",
    "CPU usage is 70% or higher.": "CPU 사용률 70% 이상",
    "CPU usage is between 50% and 69%.": "CPU 사용률 50~69%",
    "CPU usage is below 50%.": "CPU 사용률 50% 미만",
    "Memory FreeRatio is 30% or lower.": "메모리 FreeRatio 30% 이하",
    "Memory FreeRatio is between 31% and 40%.": "메모리 FreeRatio 31~40%",
    "Memory FreeRatio is above 40%.": "메모리 FreeRatio 40% 초과",
    "Power State is not Normal.": "전원 State 비정상",
}

CHANGE_TYPE_LABELS_KO = {
    "changed": "변경",
    "added": "추가",
    "removed": "삭제",
    "context": "문맥",
    "unchanged": "변경없음",
}

EXPECTATION_LABELS_KO = {
    "unexpected": "문제",
    "expected": "예상된 변화",
    "unknown": "참고",
}

SEVERITY_SORT_ORDER = {"Critical": 0, "Warning": 1, "Info": 2, "Unchanged": 3}
EXPECTATION_SORT_ORDER = {"unexpected": 0, "expected": 1, "unknown": 2}

SEVERITY_COLORS = {
    "Critical": "#b42318",
    "Warning": "#b54708",
    "Info": "#2563a8",
    "Unchanged": "#007a63",
}
SEVERITY_SOFT_COLORS = {
    "Critical": "#fdeceb",
    "Warning": "#fff4e5",
    "Info": "#e8f1fb",
    "Unchanged": "#ddf5f0",
}


class ReportWriter:
    def write_reports(self, summary: DiffSummary) -> dict[str, Path]:
        base_name = sanitize_filename(Path(summary.base_snapshot).name)
        target_dir = Path(summary.target_snapshot)
        report_dir = target_dir / "comparisons" / f"vs_{base_name}"
        report_dir.mkdir(parents=True, exist_ok=True)

        manifest_path = report_dir / "diff_manifest.json"
        html_path = report_dir / "diff_report.html"
        xlsx_path = report_dir / "diff_summary.xlsx"
        csv_path = report_dir / "diff_summary.csv"

        manifest_path.write_text(
            json.dumps(redact_payload(asdict(summary)), indent=2, ensure_ascii=True),
            encoding="utf-8",
        )
        self._write_html(html_path, summary)
        written_xlsx = self._write_xlsx(xlsx_path, summary)
        if not written_xlsx:
            self._write_csv(csv_path, summary)
        share_zip_path = create_share_report_bundle(report_dir)

        paths = {"html": html_path, "json": manifest_path}
        if written_xlsx:
            paths["xlsx"] = xlsx_path
        else:
            paths["csv"] = csv_path
        paths["share_zip"] = share_zip_path
        return paths

    @staticmethod
    def _rows(summary: DiffSummary) -> list[dict[str, str]]:
        return [
            {
                "expectation": expectation_label(item),
                "severity": item.severity,
                "status": item.status,
                "device": item.device_name,
                "command_id": item.command_id,
                "command": redact_sensitive_text(item.command),
                "category": item.category,
                "finding_title": redact_sensitive_text(finding_title(item)),
                "impact_reason": redact_sensitive_text(item.impact_reason),
                "evidence": redact_sensitive_text(item.evidence or item.change_preview),
                "action_hint": redact_sensitive_text(item.action_hint),
                "priority": str(item.priority),
                "summary": summary_label(item),
                "change_count": str(item.change_count),
                "change_preview": redact_sensitive_text(item.change_preview),
                "base_raw_file": item.base_raw_file,
                "target_raw_file": item.target_raw_file,
            }
            for item in summary.items
        ]

    @staticmethod
    def _detail_rows(summary: DiffSummary) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for item in summary.items:
            for line in item.changed_lines:
                if line.kind == "context":
                    continue
                rows.append(
                    {
                        "severity": item.severity,
                        "status": item.status,
                        "device": item.device_name,
                        "command_id": item.command_id,
                        "command": redact_sensitive_text(item.command),
                        "category": item.category,
                        "change_type": change_type_label(line.kind),
                        "base_line_no": str(line.base_line_no or ""),
                        "target_line_no": str(line.target_line_no or ""),
                        "base_text": redact_sensitive_text(line.base_text),
                        "target_text": redact_sensitive_text(line.target_text),
                    }
                )
        return rows

    def _write_csv(self, path: Path, summary: DiffSummary) -> None:
        rows = self._rows(summary)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else ["severity"])
            writer.writeheader()
            writer.writerows(rows)

    def _write_xlsx(self, path: Path, summary: DiffSummary) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:  # pragma: no cover
            return False

        summary_rows = sorted(self._rows(summary), key=lambda row: problem_row_sort_key(row))
        summary_headers = (
            list(summary_rows[0].keys())
            if summary_rows
            else [
                "expectation",
                "severity",
                "status",
                "device",
                "command_id",
                "finding_title",
                "impact_reason",
                "evidence",
                "action_hint",
                "priority",
                "summary",
                "change_count",
                "change_preview",
            ]
        )
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "problem_summary"
        summary_sheet.append(summary_headers)
        for row in summary_rows:
            summary_sheet.append([row.get(header, "") for header in summary_headers])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        severity_fill = {
            "Critical": PatternFill("solid", fgColor="F4CCCC"),
            "Warning": PatternFill("solid", fgColor="FCE5CD"),
            "Info": PatternFill("solid", fgColor="D9EAD3"),
            "Unchanged": PatternFill("solid", fgColor="EFEFEF"),
        }
        style_worksheet(summary_sheet, summary_headers, header_fill, severity_fill, max_width=70)

        detail_rows = self._detail_rows(summary)
        detail_headers = (
            list(detail_rows[0].keys())
            if detail_rows
            else [
                "severity",
                "status",
                "device",
                "command_id",
                "command",
                "category",
                "change_type",
                "base_line_no",
                "target_line_no",
                "base_text",
                "target_text",
            ]
        )
        detail_sheet = workbook.create_sheet("diff_detail")
        detail_sheet.append(detail_headers)
        for row in detail_rows:
            detail_sheet.append([row.get(header, "") for header in detail_headers])
        style_worksheet(detail_sheet, detail_headers, header_fill, severity_fill, max_width=90)

        for sheet in (summary_sheet, detail_sheet):
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        workbook.save(path)
        return True

    @staticmethod
    def _write_html(path: Path, summary: DiffSummary) -> None:
        counts = summary.counts
        base_display_name = snapshot_display_name(Path(summary.base_snapshot))
        target_display_name = snapshot_display_name(Path(summary.target_snapshot))
        status_jump_buttons_html = []
        summary_cards_html = []
        unchanged_summary_cards_html = []
        details_html = []
        problem_items = [item for item in sorted(summary.items, key=problem_item_sort_key) if is_problem_item(item)]
        for index, item in enumerate(summary.items, start=1):
            detail_id = f"diff-{index}"
            severity_label = severity_to_korean(item.severity)
            status_label = status_to_korean(item.status)
            item_summary = redact_sensitive_text(summary_label(item))
            change_preview = redact_sensitive_text(item.change_preview or "-")
            item_title = redact_sensitive_text(finding_title(item))
            impact_reason = redact_sensitive_text(item.impact_reason or "-")
            evidence = redact_sensitive_text(item.evidence or change_preview)
            action_hint = redact_sensitive_text(item.action_hint or "-")
            default_hidden_attr = " hidden aria-hidden='true'"
            card_html = (
                f"<article class='summary-card' data-severity='{escape(item.severity)}' aria-labelledby='summary-{index}'{default_hidden_attr}>"
                "<div class='summary-card-head'>"
                f"<span class='badge' style='background:{SEVERITY_COLORS.get(item.severity, '#667085')}'>{escape(severity_label)}</span>"
                f"<strong id='summary-{index}'>{escape(item.device_name)} / {escape(item.command_id)}</strong>"
                f"<a class='summary-link' data-target-severity='{escape(item.severity)}' href='#{detail_id}'>상세 보기</a>"
                "</div>"
                "<div class='summary-meta'>"
                f"<div class='summary-item'><span class='summary-label'>등급</span><span class='summary-value'>{escape(severity_label)}</span></div>"
                f"<div class='summary-item'><span class='summary-label'>상태</span><span class='summary-value'>{escape(status_label)}</span></div>"
                f"<div class='summary-item'><span class='summary-label'>장비</span><span class='summary-value'>{escape(item.device_name)}</span></div>"
                f"<div class='summary-item'><span class='summary-label'>명령</span><span class='summary-value'>{escape(item.command_id)}</span></div>"
                f"<div class='summary-item'><span class='summary-label'>분류</span><span class='summary-value'>{escape(item.category)}</span></div>"
                f"<div class='summary-item'><span class='summary-label'>변경 수</span><span class='summary-value'>{item.change_count}</span></div>"
                "</div>"
                f"<div class='summary-field'><span class='summary-label'>문제 제목</span><span class='summary-value'>{escape(item_title)}</span></div>"
                f"<div class='summary-field'><span class='summary-label'>영향 이유</span><span class='summary-value'>{escape(impact_reason)}</span></div>"
                f"<div class='summary-field'><span class='summary-label'>판단 근거</span><span class='summary-value'>{escape(evidence)}</span></div>"
                f"<div class='summary-field'><span class='summary-label'>권장 조치</span><span class='summary-value'>{escape(action_hint)}</span></div>"
                f"<div class='summary-field'><span class='summary-label'>첫 변경</span><span class='summary-value'>{escape(change_preview)}</span></div>"
                f"<div class='summary-field'><span class='summary-label'>요약</span><span class='summary-value'>{escape(item_summary)}</span></div>"
                "</article>"
            )
            status_jump_buttons_html.append(
                "<button class='jump-button' type='button' "
                f"data-jump-target='{detail_id}' data-jump-severity='{escape(item.severity)}'{default_hidden_attr}>"
                f"<span class='jump-severity' style='background:{SEVERITY_COLORS.get(item.severity, '#667085')}'>{escape(severity_label)}</span>"
                f"<span class='jump-main'>{escape(item.device_name)} / {escape(item.command_id)}</span>"
                f"<span class='jump-count'>{item.change_count}건</span>"
                "</button>"
            )
            if item.severity == "Unchanged":
                unchanged_summary_cards_html.append(card_html)
            else:
                summary_cards_html.append(card_html)
            details_html.append(render_diff_block(item, detail_id, severity_label, status_label, item_summary))

        status_jump_html = (
            "<section class='jump-list' aria-label='상태별 바로가기' data-jump-list hidden aria-hidden='true'>"
            "<div class='jump-head'>"
            "<strong>상태별 바로가기</strong>"
            "<span class='meta'>상단 상태 카드를 클릭하면 선택한 상태의 장비/명령 버튼만 표시합니다.</span>"
            "</div>"
            f"<div class='jump-actions'>{''.join(status_jump_buttons_html)}</div>"
            "<p class='meta jump-empty' data-jump-empty hidden aria-hidden='true'>선택한 상태의 바로가기 항목 없음</p>"
            "</section>"
        )
        visible_summary_html = (
            "".join(summary_cards_html)
            + "<p class='meta summary-empty' data-summary-empty hidden aria-hidden='true'>선택한 상태의 요약 항목 없음</p>"
        )
        unchanged_summary_html = ""
        if unchanged_summary_cards_html:
            unchanged_count = counts.get("Unchanged", 0)
            unchanged_summary_html = (
                "<details class='summary-list unchanged-summary' data-summary-severity='Unchanged' hidden aria-hidden='true'>"
                "<summary>"
                "<span class='collapsed-head'>"
                f"<span class='badge' style='background:{SEVERITY_COLORS['Unchanged']}'>변경없음</span>"
                f"<span>변경없음 {unchanged_count}건 - 필요 시 펼쳐서 확인</span>"
                "</span>"
                "</summary>"
                f"<div class='unchanged-summary-body'>{''.join(unchanged_summary_cards_html)}</div>"
                "</details>"
            )
        problem_summary_html = render_problem_summary(problem_items)

        html = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(APP_NAME)} v{escape(APP_VERSION)} 스냅샷 비교 리포트</title>
  <style>
    :root {{
      --bg: #eef3f6;
      --panel: #ffffff;
      --panel-alt: #f7fafb;
      --line: #ced9e2;
      --line-strong: #aab7c2;
      --text: #17212b;
      --muted: #60717e;
      --accent: #01a982;
      --accent-dark: #007a63;
      --accent-soft: #ddf5f0;
      --added: #ddf5f0;
      --removed: #fdeceb;
      --changed: #fff4e5;
      --context: #f7fafb;
{css_severity_vars()}
    }}
    * {{ box-sizing: border-box; }}
    [hidden] {{ display: none !important; }}
    body {{
      margin: 0;
      font-family: "Malgun Gothic", "Segoe UI", Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.45;
    }}
    .wrap {{ max-width: 1360px; margin: 0 auto; padding: 24px; }}
    header {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-left: 5px solid var(--accent);
      border-radius: 8px;
      margin-bottom: 18px;
      padding: 18px 20px;
    }}
    h1 {{ margin: 0 0 8px; font-size: 26px; }}
    .meta {{ color: var(--muted); font-size: 13px; }}
    .counts {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin: 18px 0;
    }}
    .filter-state {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-left: 5px solid var(--accent);
      border-radius: 8px;
      color: var(--muted);
      font-size: 13px;
      margin: 0 0 14px;
      padding: 11px 14px;
      box-shadow: 0 1px 2px rgba(16, 24, 32, 0.04);
    }}
    .count {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-left-width: 5px;
      border-radius: 8px;
      padding: 14px;
      cursor: pointer;
      color: var(--text);
      font: inherit;
      text-align: left;
      box-shadow: 0 1px 2px rgba(16, 24, 32, 0.04);
    }}
    .count[data-filter="Critical"] {{ border-left-color: var(--critical); background: linear-gradient(90deg, var(--critical-soft), var(--panel) 42%); }}
    .count[data-filter="Warning"] {{ border-left-color: var(--warning); background: linear-gradient(90deg, var(--warning-soft), var(--panel) 42%); }}
    .count[data-filter="Info"] {{ border-left-color: var(--info); background: linear-gradient(90deg, var(--info-soft), var(--panel) 42%); }}
    .count[data-filter="Unchanged"] {{ border-left-color: var(--unchanged); background: linear-gradient(90deg, var(--unchanged-soft), var(--panel) 42%); }}
    .count-label {{ display: block; color: var(--muted); font-size: 12px; font-weight: 800; }}
    .count strong {{ display: block; font-size: 24px; margin-top: 4px; }}
    .count-hint {{ display: block; color: var(--muted); font-size: 12px; margin-top: 2px; }}
    .count:hover {{ border-color: var(--accent); }}
    .count.is-active {{ outline: 2px solid var(--accent); border-color: var(--accent); }}
    .problem-summary {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-left: 5px solid var(--critical);
      border-radius: 8px;
      margin: 0 0 16px;
      padding: 16px;
      box-shadow: 0 1px 2px rgba(16, 24, 32, 0.04);
    }}
    .problem-summary h2 {{ margin: 0 0 4px; font-size: 18px; }}
    .problem-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(290px, 1fr));
      gap: 10px;
      margin-top: 12px;
    }}
    .problem-card {{
      border: 1px solid #e3e8ee;
      border-radius: 8px;
      background: var(--panel-alt);
      padding: 12px;
    }}
    .problem-card-head {{
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
      margin-bottom: 8px;
    }}
    .problem-card h3 {{
      margin: 0 0 8px;
      font-size: 15px;
      overflow-wrap: anywhere;
    }}
    .problem-field {{ margin-top: 8px; }}
    .jump-list {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      margin: 0 0 16px;
      box-shadow: 0 1px 2px rgba(16, 24, 32, 0.04);
    }}
    .jump-head {{
      display: flex;
      align-items: baseline;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 10px;
    }}
    .jump-head strong {{ font-size: 15px; }}
    .jump-actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}
    .jump-button {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      max-width: 100%;
      border: 1px solid #d0d7de;
      border-radius: 999px;
      background: #fff;
      color: var(--text);
      cursor: pointer;
      font: inherit;
      font-size: 13px;
      padding: 6px 10px;
      text-align: left;
    }}
    .jump-button:hover {{ border-color: var(--accent); box-shadow: 0 0 0 2px rgba(1, 169, 130, 0.12); }}
    .jump-severity {{
      color: #fff;
      border-radius: 999px;
      padding: 2px 7px;
      font-size: 12px;
      font-weight: 700;
      flex: 0 0 auto;
    }}
    .jump-main {{ overflow-wrap: anywhere; }}
    .jump-count {{ color: var(--muted); font-size: 12px; flex: 0 0 auto; }}
    .summary-list, .diff-block {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      margin-bottom: 16px;
      box-shadow: 0 1px 2px rgba(16, 24, 32, 0.04);
    }}
    .summary-list {{ padding: 12px; }}
    .unchanged-summary {{ padding: 0; }}
    .unchanged-summary > summary {{
      cursor: pointer;
      list-style: none;
      padding: 14px;
    }}
    .unchanged-summary > summary::-webkit-details-marker {{ display: none; }}
    .unchanged-summary-body {{ padding: 0 12px 12px; }}
    .summary-card {{
      border: 1px solid #e3e8ee;
      border-radius: 8px;
      padding: 14px;
      background: #fff;
    }}
    .summary-card + .summary-card {{ margin-top: 10px; }}
    .summary-card-head {{
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 12px;
    }}
    .summary-card-head strong {{
      font-size: 15px;
      overflow-wrap: anywhere;
    }}
    .summary-link {{
      margin-left: auto;
      font-size: 13px;
      font-weight: 700;
    }}
    .summary-meta {{
      display: grid;
      grid-template-columns: repeat(6, minmax(110px, 1fr));
      gap: 8px;
      margin-bottom: 8px;
    }}
    .summary-item, .summary-field {{
      background: var(--panel-alt);
      border: 1px solid #e5ebf0;
      border-radius: 6px;
      padding: 8px 10px;
    }}
    .summary-field {{
      display: grid;
      grid-template-columns: 90px minmax(0, 1fr);
      gap: 10px;
      margin-top: 8px;
    }}
    .summary-label {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      margin-bottom: 3px;
    }}
    .summary-field .summary-label {{ margin-bottom: 0; }}
    .summary-value {{
      display: block;
      font-size: 13px;
      line-height: 1.5;
      overflow-wrap: anywhere;
      word-break: break-word;
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 10px; border-bottom: 1px solid #eaeef2; text-align: left; font-size: 13px; vertical-align: top; }}
    th {{ background: #e6edf2; color: #17212b; }}
    .badge {{ color: #fff; border-radius: 999px; padding: 3px 8px; font-size: 12px; font-weight: 700; }}
    .diff-block {{
      padding: 14px;
      border-top-width: 4px;
    }}
    .diff-block[data-severity="Critical"] {{ border-top-color: var(--critical); }}
    .diff-block[data-severity="Warning"] {{ border-top-color: var(--warning); }}
    .diff-block[data-severity="Info"] {{ border-top-color: var(--info); }}
    .diff-block[data-severity="Unchanged"] {{ border-top-color: var(--unchanged); }}
    .diff-block.is-jump-target {{ outline: 3px solid rgba(1, 169, 130, 0.32); border-color: var(--accent); }}
    .detail-head {{
      display: grid;
      grid-template-columns: auto minmax(0, 1fr);
      gap: 10px;
      align-items: start;
      margin-bottom: 10px;
    }}
    .detail-title h2 {{ margin: 0 0 4px; font-size: 17px; overflow-wrap: anywhere; }}
    .detail-meta {{ color: var(--muted); font-size: 13px; }}
    details.diff-block > summary {{
      cursor: pointer;
      list-style: none;
    }}
    details.diff-block > summary::-webkit-details-marker {{ display: none; }}
    .collapsed-head {{
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
      font-weight: 700;
    }}
    .change-table-wrap {{
      overflow-x: auto;
      margin: 12px 0;
      border: 1px solid #e5ebf0;
      border-radius: 6px;
    }}
    .change-table {{ width: max-content; min-width: 100%; border-collapse: collapse; margin: 0; }}
    .change-table th:nth-child(1) {{ width: 82px; }}
    .change-table th:nth-child(2) {{ width: 110px; }}
    .change-table td {{ word-break: normal; white-space: nowrap; }}
    .change-kind {{
      display: inline-flex;
      align-items: center;
      border: 1px solid #d0d7de;
      border-radius: 999px;
      padding: 2px 8px;
      background: #fff;
      font-size: 12px;
      font-weight: 700;
      color: #1f2328;
    }}
    .cell-label {{
      display: inline-block;
      margin-right: 6px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }}
    .line-no {{
      color: var(--muted);
      font-variant-numeric: tabular-nums;
    }}
    .line-value {{
      color: var(--text);
    }}
    .change-inline {{
      font-family: Consolas, "SFMono-Regular", monospace;
      line-height: 1.7;
    }}
    .change-inline .cell-label {{
      font-family: "Malgun Gothic", "Segoe UI", Arial, sans-serif;
    }}
    .value-before, .value-after, .value-context {{
      display: inline-block;
      border: 1px solid transparent;
      border-radius: 4px;
      padding: 2px 6px;
    }}
    .value-before {{
      background: var(--removed);
      border-color: #f4b8b2;
      color: #8a1f16;
    }}
    .value-after {{
      background: var(--added);
      border-color: #9bd8cb;
      color: var(--accent-dark);
    }}
    .value-context {{
      color: var(--muted);
      padding-left: 0;
    }}
    .diff-arrow {{ color: var(--muted); font-weight: 700; padding: 0 8px; }}
    .diff-prefix {{ color: var(--muted); font-weight: 700; margin-right: 6px; }}
    .line-added {{ background: var(--added); }}
    .line-removed {{ background: var(--removed); }}
    .line-changed {{ background: var(--changed); }}
    .line-context {{ background: var(--context); color: var(--muted); }}
    .raw-diff summary {{ cursor: pointer; color: #0969da; font-weight: 700; margin: 10px 0; }}
    pre {{
      background: #101820;
      color: #dce6ed;
      border-radius: 6px;
      padding: 12px;
      overflow: auto;
      line-height: 1.4;
      font-size: 12px;
    }}
    a {{ color: #0969da; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    @media (max-width: 900px) {{
      .counts {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .summary-meta {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .summary-field {{ display: block; }}
      .summary-link {{ margin-left: 0; }}
      .jump-head {{ display: block; }}
      .jump-button {{ width: 100%; border-radius: 8px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <header>
      <h1>{escape(APP_NAME)} 스냅샷 비교 리포트</h1>
      <div class="meta">버전: v{escape(APP_VERSION)} | 기준: {escape(base_display_name)} | 비교: {escape(target_display_name)} | 생성: {escape(summary.generated_at)}</div>
    </header>
    <section class="filter-state" data-filter-state>상태 카드를 선택하면 해당 등급의 바로가기, 요약, 상세만 표시합니다.</section>
    <section class="counts" aria-label="등급 필터">
      <button class="count" type="button" data-filter="Critical"><span class="count-label">긴급</span><strong>{counts.get('Critical', 0)}</strong><span class="count-hint">즉시 확인</span></button>
      <button class="count" type="button" data-filter="Warning"><span class="count-label">주의</span><strong>{counts.get('Warning', 0)}</strong><span class="count-hint">영향 확인</span></button>
      <button class="count" type="button" data-filter="Info"><span class="count-label">정보</span><strong>{counts.get('Info', 0)}</strong><span class="count-hint">기록 확인</span></button>
      <button class="count" type="button" data-filter="Unchanged"><span class="count-label">변경없음</span><strong>{counts.get('Unchanged', 0)}</strong><span class="count-hint">필요 시 펼침</span></button>
    </section>
    {problem_summary_html}
    {status_jump_html}
    <section class="summary-list" aria-label="비교 요약" data-summary-list hidden aria-hidden="true">
      {visible_summary_html}
    </section>
    {unchanged_summary_html}
    {''.join(details_html)}
  </div>
  <script>
    const filterButtons = Array.from(document.querySelectorAll("[data-filter]"));
    const filterEntries = Array.from(document.querySelectorAll(".diff-block[data-severity]"));
    const summaryCards = Array.from(document.querySelectorAll(".summary-card[data-severity]"));
    const summaryLinks = Array.from(document.querySelectorAll("[data-target-severity]"));
    const jumpButtons = Array.from(document.querySelectorAll("[data-jump-target]"));
    const jumpList = document.querySelector("[data-jump-list]");
    const summaryList = document.querySelector("[data-summary-list]");
    const unchangedSummary = document.querySelector("[data-summary-severity='Unchanged']");
    const filterState = document.querySelector("[data-filter-state]");
    const summaryEmpty = document.querySelector("[data-summary-empty]");
    const jumpEmpty = document.querySelector("[data-jump-empty]");
    const mainSummaryCards = summaryCards.filter((card) => !card.closest("[data-summary-severity='Unchanged']"));
    const filterLabels = {{"Critical": "긴급", "Warning": "주의", "Info": "정보", "Unchanged": "변경없음"}};
    let activeFilter = "";
    function severityVisible(severity) {{
      return activeFilter ? severity === activeFilter : false;
    }}
    function setElementHidden(element, hidden) {{
      if (!element) return;
      element.hidden = hidden;
      element.setAttribute("aria-hidden", hidden ? "true" : "false");
    }}
    function focusTarget(target) {{
      if (!target) return;
      if (target.tagName.toLowerCase() === "details") {{
        target.open = true;
      }}
      target.classList.add("is-jump-target");
      target.scrollIntoView({{ behavior: "smooth", block: "start" }});
      target.focus({{ preventScroll: true }});
      window.setTimeout(() => target.classList.remove("is-jump-target"), 1800);
    }}
    function setFilter(nextFilter) {{
      activeFilter = nextFilter;
      filterButtons.forEach((button) => {{
        button.classList.toggle("is-active", button.dataset.filter === activeFilter);
      }});
      if (filterState) {{
        filterState.textContent = activeFilter
          ? `${{filterLabels[activeFilter] || activeFilter}} 상태만 표시 중입니다. 상태 카드를 다시 누르면 목록을 숨깁니다.`
          : "상태 카드를 선택하면 해당 등급의 바로가기, 요약, 상세만 표시합니다.";
      }}
      filterEntries.forEach((entry) => {{
        const visible = severityVisible(entry.dataset.severity);
        setElementHidden(entry, !visible);
      }});
      summaryCards.forEach((card) => {{
        const visible = severityVisible(card.dataset.severity);
        setElementHidden(card, !visible);
      }});
      jumpButtons.forEach((button) => {{
        const visible = severityVisible(button.dataset.jumpSeverity);
        setElementHidden(button, !visible);
      }});
      setElementHidden(jumpList, !activeFilter);
      setElementHidden(summaryList, !activeFilter || activeFilter === "Unchanged");
      if (summaryEmpty) {{
        const hasVisibleMainSummary = mainSummaryCards.some((card) => !card.hidden);
        setElementHidden(summaryEmpty, !activeFilter || activeFilter === "Unchanged" || hasVisibleMainSummary);
      }}
      if (jumpEmpty) {{
        const hasVisibleJump = jumpButtons.some((button) => !button.hidden);
        setElementHidden(jumpEmpty, !activeFilter || hasVisibleJump);
      }}
      if (unchangedSummary) {{
        setElementHidden(unchangedSummary, !severityVisible("Unchanged"));
        if (activeFilter !== "Unchanged") {{
          unchangedSummary.open = false;
        }}
      }}
    }}
    function applyFilter(nextFilter) {{
      setFilter(activeFilter === nextFilter ? "" : nextFilter);
    }}
    filterButtons.forEach((button) => {{
      button.addEventListener("click", () => applyFilter(button.dataset.filter));
    }});
    jumpButtons.forEach((button) => {{
      button.addEventListener("click", () => {{
        setFilter(button.dataset.jumpSeverity);
        const target = document.getElementById(button.dataset.jumpTarget);
        focusTarget(target);
      }});
    }});
    summaryLinks.forEach((link) => {{
      link.addEventListener("click", () => {{
        if (link.dataset.targetSeverity === "Unchanged" && unchangedSummary) {{
          unchangedSummary.open = true;
        }}
        if (activeFilter && activeFilter !== link.dataset.targetSeverity) {{
          setFilter(link.dataset.targetSeverity);
        }}
        const targetId = link.getAttribute("href");
        if (targetId && targetId.startsWith("#")) {{
          const target = document.querySelector(targetId);
          focusTarget(target);
        }}
      }});
    }});
    setFilter("");
  </script>
</body>
</html>
"""
        path.write_text(html, encoding="utf-8")


def style_worksheet(sheet, headers, header_fill, severity_fill, max_width: int) -> None:
    from openpyxl.styles import Font

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if "severity" in headers and sheet.max_row > 1:
        severity_col = headers.index("severity") + 1
        for row in range(2, sheet.max_row + 1):
            severity = sheet.cell(row=row, column=severity_col).value
            fill = severity_fill.get(severity)
            if fill:
                sheet.cell(row=row, column=severity_col).fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), max_width)


def summary_label(item: DiffItem) -> str:
    return SUMMARY_LABELS_KO.get(item.summary, item.summary)


def finding_title(item: DiffItem) -> str:
    return item.finding_title or summary_label(item)


def expectation_label(item: DiffItem) -> str:
    return EXPECTATION_LABELS_KO.get(item.expectation, "참고")


def is_problem_item(item: DiffItem) -> bool:
    return item.expectation == "unexpected" and item.severity in {"Critical", "Warning"}


def problem_item_sort_key(item: DiffItem) -> tuple[int, int, int, str, str]:
    return (
        EXPECTATION_SORT_ORDER.get(item.expectation, 2),
        SEVERITY_SORT_ORDER.get(item.severity, 9),
        item.priority,
        item.device_name,
        item.command_id,
    )


def problem_row_sort_key(row: dict[str, str]) -> tuple[int, int, int, str, str]:
    expectation_order = {"문제": 0, "예상된 변화": 1, "참고": 2}
    try:
        priority = int(row.get("priority", "50"))
    except ValueError:
        priority = 50
    return (
        expectation_order.get(row.get("expectation", "참고"), 2),
        SEVERITY_SORT_ORDER.get(row.get("severity", ""), 9),
        priority,
        row.get("device", ""),
        row.get("command_id", ""),
    )


def change_type_label(kind: str) -> str:
    return CHANGE_TYPE_LABELS_KO.get(kind, kind)


def css_severity_vars() -> str:
    names = ("Critical", "Warning", "Info", "Unchanged")
    color_vars = [f"      --{name.lower()}: {SEVERITY_COLORS[name]};" for name in names]
    soft_vars = [f"      --{name.lower()}-soft: {SEVERITY_SOFT_COLORS[name]};" for name in names]
    return "\n".join(color_vars + soft_vars)


def snapshot_display_name(snapshot_dir: Path) -> str:
    name = snapshot_dir.name
    try:
        snapshot = SnapshotStore.load_snapshot(snapshot_dir)
    except Exception:
        return name
    if is_sample_snapshot(snapshot_dir, snapshot.stage_slug):
        return f"샘플: {name}"
    return name


def render_problem_summary(problem_items: list[DiffItem]) -> str:
    if not problem_items:
        return (
            "<section class='problem-summary' aria-label='먼저 확인할 문제'>"
            "<h2>먼저 확인할 문제</h2>"
            "<p class='meta'>예상되지 않은 긴급/주의 문제가 없습니다.</p>"
            "</section>"
        )

    cards = "".join(render_problem_card(item) for item in problem_items)
    return (
        "<section class='problem-summary' aria-label='먼저 확인할 문제'>"
        "<h2>먼저 확인할 문제</h2>"
        "<p class='meta'>예상되지 않은 긴급/주의 항목만 먼저 표시합니다. 전체 변경 라인은 아래 상태 필터에서 확인할 수 있습니다.</p>"
        f"<div class='problem-grid'>{cards}</div>"
        "</section>"
    )


def render_problem_card(item: DiffItem) -> str:
    severity_label = severity_to_korean(item.severity)
    title = redact_sensitive_text(finding_title(item))
    evidence = redact_sensitive_text(item.evidence or item.change_preview or summary_label(item))
    action_hint = redact_sensitive_text(item.action_hint or "-")
    return (
        "<article class='problem-card'>"
        "<div class='problem-card-head'>"
        f"<span class='badge' style='background:{SEVERITY_COLORS.get(item.severity, '#667085')}'>{escape(severity_label)}</span>"
        f"<span class='meta'>{escape(item.device_name)} / {escape(item.command_id)}</span>"
        "</div>"
        f"<h3>{escape(title)}</h3>"
        f"<div class='problem-field'><span class='summary-label'>판단 근거</span><span class='summary-value'>{escape(evidence)}</span></div>"
        f"<div class='problem-field'><span class='summary-label'>권장 조치</span><span class='summary-value'>{escape(action_hint)}</span></div>"
        "</article>"
    )


def render_diff_block(item: DiffItem, detail_id: str, severity_label: str, status_label: str, item_summary: str) -> str:
    default_hidden_attr = " hidden aria-hidden='true'"
    body = (
        f"{render_interpretation(item)}"
        f"<p>{escape(item_summary)}</p>"
        f"{render_change_table(item)}"
        "<details class='raw-diff'>"
        "<summary>원본 unified diff 보기</summary>"
        f"<pre>{escape(redact_sensitive_text(item.diff or '상세 diff 없음'))}</pre>"
        "</details>"
    )
    if item.severity == "Unchanged":
        return (
            f"<details class='diff-block filter-entry' id='{detail_id}' data-severity='{escape(item.severity)}' tabindex='-1'{default_hidden_attr}>"
            "<summary>"
            "<span class='collapsed-head'>"
            f"<span class='badge' style='background:{SEVERITY_COLORS['Unchanged']}'>변경없음</span>"
            f"<span>{escape(item.device_name)} / {escape(item.command_id)}</span>"
            f"<span class='meta'>등급: {escape(severity_label)} | 상태: {escape(status_label)} | 분류: {escape(item.category)} | 변경 수: {item.change_count}</span>"
            "</span>"
            "</summary>"
            f"{body}"
            "</details>"
        )
    return (
        f"<section class='diff-block filter-entry' id='{detail_id}' data-severity='{escape(item.severity)}' tabindex='-1'{default_hidden_attr}>"
        f"{render_detail_head(item, severity_label, status_label)}"
        f"{body}"
        "</section>"
    )


def render_detail_head(item: DiffItem, severity_label: str, status_label: str) -> str:
    return (
        "<div class='detail-head'>"
        f"<span class='badge' style='background:{SEVERITY_COLORS.get(item.severity, '#667085')}'>{escape(severity_label)}</span>"
        "<div class='detail-title'>"
        f"<h2>{escape(item.device_name)} / {escape(item.command_id)}</h2>"
        f"<div class='detail-meta'>상태: {escape(status_label)} | 분류: {escape(item.category)} | 변경 수: {item.change_count}</div>"
        "</div>"
        "</div>"
    )


def render_interpretation(item: DiffItem) -> str:
    title = redact_sensitive_text(finding_title(item))
    impact_reason = redact_sensitive_text(item.impact_reason or "기준과 비교 시점의 상태가 달라졌습니다.")
    evidence = redact_sensitive_text(item.evidence or item.change_preview or summary_label(item))
    action_hint = redact_sensitive_text(item.action_hint or "-")
    return (
        "<div class='summary-field'><span class='summary-label'>문제 제목</span><span class='summary-value'>"
        f"{escape(title)}</span></div>"
        "<div class='summary-field'><span class='summary-label'>영향 이유</span><span class='summary-value'>"
        f"{escape(impact_reason)}</span></div>"
        "<div class='summary-field'><span class='summary-label'>판단 근거</span><span class='summary-value'>"
        f"{escape(evidence)}</span></div>"
        "<div class='summary-field'><span class='summary-label'>권장 조치</span><span class='summary-value'>"
        f"{escape(action_hint)}</span></div>"
    )


def render_change_table(item: DiffItem) -> str:
    if not item.changed_lines:
        return "<p class='meta'>행 단위 변경 상세 없음</p>"

    rows = []
    for line in item.changed_lines:
        rows.append(
            f"<tr class='{line_css_class(line)}'>"
            "<td data-label='유형'><span class='cell-label'>유형</span>"
            f"<span class='change-kind'>{escape(change_type_label(line.kind))}</span></td>"
            "<td class='line-no' data-label='라인'><span class='cell-label'>라인</span>"
            f"<span class='line-value'>{escape(format_inline_line_no(line))}</span></td>"
            "<td class='change-inline' data-label='변경 내용'><span class='cell-label'>변경 내용</span>"
            f"{render_inline_change(line)}</td>"
            "</tr>"
        )
    return (
        "<div class='change-table-wrap'>"
        "<table class='change-table'>"
        "<thead><tr><th>유형</th><th>라인</th><th>변경 내용</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
        "</div>"
    )


def line_css_class(line: DiffLine) -> str:
    return {
        "added": "line-added",
        "removed": "line-removed",
        "changed": "line-changed",
        "context": "line-context",
    }.get(line.kind, "")


def format_line_no(value: int | None) -> str:
    return "-" if value is None else str(value)


def format_inline_line_no(line: DiffLine) -> str:
    base_line_no = format_line_no(line.base_line_no)
    target_line_no = format_line_no(line.target_line_no)
    if line.kind == "added":
        return f"- → {target_line_no}"
    if line.kind == "removed":
        return f"{base_line_no} → -"
    return f"{base_line_no} → {target_line_no}"


def render_inline_change(line: DiffLine) -> str:
    if line.kind == "changed":
        return (
            f"{render_value(line.base_text, 'value-before')}"
            "<span class='diff-arrow'>→</span>"
            f"{render_value(line.target_text, 'value-after')}"
        )
    if line.kind == "added":
        return f"<span class='diff-prefix'>추가:</span>{render_value(line.target_text, 'value-after')}"
    if line.kind == "removed":
        return f"<span class='diff-prefix'>삭제:</span>{render_value(line.base_text, 'value-before')}"
    return render_value(line.target_text or line.base_text, "value-context")


def render_value(value: str, css_class: str) -> str:
    return f"<span class='{css_class}'>{escape(redact_sensitive_text(value) or '-')}</span>"
