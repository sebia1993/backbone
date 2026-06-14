from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

from backbone_state_tracker.core.diff_engine import DiffEngine
from backbone_state_tracker.core.models import CommandResult, Device, DiffItem, DiffLine, DiffSummary
from backbone_state_tracker.core.reporter import ReportWriter
from backbone_state_tracker.core.snapshot import SnapshotStore


class ReportWriterTests(unittest.TestCase):
    def _snapshot(
        self,
        root: Path,
        label: str,
        output: str,
        *,
        folder_label: str | None = None,
        stage_name: str = "",
        stage_slug: str = "",
    ) -> Path:
        store = SnapshotStore(root)
        device = Device(name="backbone4", host="192.0.2.4")
        result = CommandResult(
            device_name=device.name,
            host=device.host,
            command_id="interface_brief",
            command="display interface brief",
            description="Interface summary",
            category="interface",
            phase="check",
            success=True,
            output=output,
            started_at="2026-06-11T10:00:00",
            ended_at="2026-06-11T10:00:01",
        )
        return store.write_snapshot(
            label,
            [device],
            {device.name: [result]},
            folder_label=folder_label,
            stage_name=stage_name,
            stage_slug=stage_slug,
        )

    def test_html_and_xlsx_include_line_level_diff_details(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = self._snapshot(root, "base", "GE1/0/1 UP\nGE1/0/2 UP")
            target = self._snapshot(root, "target", "GE1/0/1 DOWN\nGE1/0/2 UP")
            summary = DiffEngine().compare(base, target)

            paths = ReportWriter().write_reports(summary)

            html = paths["html"].read_text(encoding="utf-8")
            self.assertIn("summary-list", html)
            self.assertIn("summary-card", html)
            self.assertIn('data-filter="Critical"', html)
            self.assertIn('data-filter="Unchanged"', html)
            self.assertIn('data-severity=', html)
            self.assertIn('data-target-severity=', html)
            self.assertIn("상태별 바로가기", html)
            self.assertNotIn("변경 항목 바로가기", html)
            self.assertIn("data-jump-target='diff-", html)
            self.assertIn("data-jump-severity='Critical'", html)
            self.assertIn("<span class='jump-main'>backbone4 / interface_brief</span>", html)
            self.assertIn("<span class='jump-count'>1건</span>", html)
            self.assertIn("<span class='jump-main'>backbone4 / device_connectivity</span>", html)
            self.assertIn("data-jump-severity='Critical' hidden", html)
            self.assertIn("data-jump-severity='Unchanged' hidden", html)
            self.assertIn("[hidden] { display: none !important; }", html)
            self.assertIn("<section class='jump-list' aria-label='상태별 바로가기' data-jump-list hidden aria-hidden='true'>", html)
            self.assertIn('<section class="summary-list" aria-label="비교 요약" data-summary-list hidden aria-hidden="true">', html)
            self.assertIn('document.querySelectorAll(".diff-block[data-severity]")', html)
            self.assertIn('document.querySelectorAll(".summary-card[data-severity]")', html)
            self.assertIn("function applyFilter", html)
            self.assertIn("function setFilter", html)
            self.assertIn("function setElementHidden", html)
            self.assertNotIn("defaultVisibleSeverities", html)
            self.assertIn("return activeFilter ? severity === activeFilter : false;", html)
            self.assertIn("setElementHidden(button, !visible);", html)
            self.assertIn("setElementHidden(card, !visible);", html)
            self.assertIn("setElementHidden(jumpList, !activeFilter);", html)
            self.assertIn('setElementHidden(summaryList, !activeFilter || activeFilter === "Unchanged");', html)
            self.assertIn("function focusTarget", html)
            self.assertIn("target.scrollIntoView", html)
            self.assertIn("target.focus", html)
            self.assertIn("setFilter(button.dataset.jumpSeverity);", html)
            self.assertNotIn("class='summary-card filter-entry'", html)
            self.assertIn("<span class='summary-label'>변경 수</span>", html)
            self.assertIn("<span class='summary-label'>첫 변경</span>", html)
            self.assertIn("<span class='summary-label'>요약</span>", html)
            self.assertNotIn("<th>첫 변경 내용</th>", html)
            self.assertIn("변경 내용", html)
            self.assertIn("change-inline", html)
            self.assertIn("1 → 1", html)
            self.assertIn("GE1/0/1 UP</span><span class='diff-arrow'>→</span><span class='value-after'>GE1/0/1 DOWN", html)
            self.assertNotIn("<th>변경 전</th>", html)
            self.assertNotIn("<th>변경 후</th>", html)
            self.assertIn("GE1/0/1 UP", html)
            self.assertIn("GE1/0/1 DOWN", html)

            if "xlsx" not in paths:
                self.skipTest("openpyxl is not installed")

            from openpyxl import load_workbook

            workbook = load_workbook(paths["xlsx"])
            self.assertIn("diff_detail", workbook.sheetnames)
            detail_sheet = workbook["diff_detail"]
            rows = list(detail_sheet.iter_rows(values_only=True))
            self.assertIn("base_text", rows[0])
            self.assertTrue(any("GE1/0/1 UP" in row for row in rows[1:]))
            self.assertTrue(any("GE1/0/1 DOWN" in row for row in rows[1:]))

    def test_html_collapses_unchanged_detail_blocks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = self._snapshot(root, "base", "GE1/0/1 UP")
            target = self._snapshot(root, "target", "GE1/0/1 UP")
            summary = DiffEngine().compare(base, target)

            paths = ReportWriter().write_reports(summary)

            html = paths["html"].read_text(encoding="utf-8")
            self.assertIn("<details class='diff-block filter-entry'", html)
            self.assertIn("data-severity='Unchanged'", html)
            self.assertIn("collapsed-head", html)
            self.assertNotIn("변경 항목 없음", html)
            self.assertIn("상태별 바로가기", html)
            self.assertIn("data-jump-target=", html)
            self.assertIn("data-jump-severity='Unchanged' hidden", html)
            self.assertIn("<details class='summary-list unchanged-summary' data-summary-severity='Unchanged' hidden aria-hidden='true'>", html)
            self.assertIn("변경없음 2건 - 필요 시 펼쳐서 확인", html)
            self.assertIn("<div class='unchanged-summary-body'>", html)
            self.assertIn("const unchangedSummary", html)
            self.assertIn('if (activeFilter !== "Unchanged")', html)
            self.assertIn("unchangedSummary.open = false;", html)
            self.assertNotIn('unchangedSummary.open = activeFilter === "Unchanged";', html)
            self.assertIn('setElementHidden(unchangedSummary, !severityVisible("Unchanged"));', html)
            self.assertIn("target.open = true;", html)
            self.assertNotIn("<details class='summary-list unchanged-summary' data-summary-severity='Unchanged' open", html)

    def test_html_status_shortcuts_filter_all_severities_and_hide_until_selected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            report_path = Path(tmp) / "report.html"
            summary = DiffSummary(
                base_snapshot="base",
                target_snapshot="target",
                generated_at="2026-06-12T10:00:00",
                items=[
                    DiffItem(
                        device_name="backbone3",
                        command_id="critical_check",
                        command="display interface brief",
                        category="interface",
                        severity="Critical",
                        status="changed",
                        summary="Critical state keyword detected in changed output.",
                        changed_lines=[DiffLine(kind="changed", base_line_no=1, target_line_no=1, base_text="UP", target_text="DOWN")],
                        change_count=1,
                        change_preview="UP -> DOWN",
                    ),
                    DiffItem(
                        device_name="backbone4",
                        command_id="warning_check",
                        command="display logbuffer",
                        category="log",
                        severity="Warning",
                        status="changed",
                        summary="Warning keyword detected in changed output.",
                        changed_lines=[DiffLine(kind="added", target_line_no=1, target_text="warning")],
                        change_count=1,
                        change_preview="warning",
                    ),
                    DiffItem(
                        device_name="backbone4",
                        command_id="info_check",
                        command="display device",
                        category="hardware",
                        severity="Info",
                        status="changed",
                        summary="Output changed.",
                        changed_lines=[DiffLine(kind="added", target_line_no=1, target_text="reachable")],
                        change_count=1,
                        change_preview="reachable",
                    ),
                    DiffItem(
                        device_name="backbone4",
                        command_id="unchanged_check",
                        command="display ospf peer",
                        category="routing",
                        severity="Unchanged",
                        status="unchanged",
                        summary="No meaningful change detected.",
                    ),
                ],
            )

            ReportWriter._write_html(report_path, summary)

            html = report_path.read_text(encoding="utf-8")
            self.assertIn("상태별 바로가기", html)
            for severity in ("Critical", "Warning", "Info", "Unchanged"):
                self.assertIn(f"data-jump-severity='{severity}'", html)
                self.assertIn(f"data-filter=\"{severity}\"", html)
                self.assertIn(f"data-jump-severity='{severity}' hidden", html)
            self.assertIn("[hidden] { display: none !important; }", html)
            self.assertIn("<article class='summary-card' data-severity='Critical' aria-labelledby='summary-1' hidden aria-hidden='true'>", html)
            self.assertIn("<article class='summary-card' data-severity='Warning' aria-labelledby='summary-2' hidden aria-hidden='true'>", html)
            self.assertIn("data-jump-severity='Info' hidden", html)
            self.assertIn("data-jump-severity='Unchanged' hidden", html)
            self.assertIn("<article class='summary-card' data-severity='Info' aria-labelledby='summary-3' hidden aria-hidden='true'>", html)
            self.assertIn("<section class='diff-block filter-entry' id='diff-1' data-severity='Critical' tabindex='-1' hidden aria-hidden='true'>", html)
            self.assertIn("<section class='diff-block filter-entry' id='diff-2' data-severity='Warning' tabindex='-1' hidden aria-hidden='true'>", html)
            self.assertIn("<section class='diff-block filter-entry' id='diff-3' data-severity='Info' tabindex='-1' hidden aria-hidden='true'>", html)
            self.assertIn("<details class='diff-block filter-entry' id='diff-4' data-severity='Unchanged' tabindex='-1' hidden aria-hidden='true'>", html)
            self.assertIn("<section class='jump-list' aria-label='상태별 바로가기' data-jump-list hidden aria-hidden='true'>", html)
            self.assertIn('<section class="summary-list" aria-label="비교 요약" data-summary-list hidden aria-hidden="true">', html)
            self.assertIn("function severityVisible", html)
            self.assertIn("function setElementHidden", html)
            self.assertIn("return activeFilter ? severity === activeFilter : false;", html)
            self.assertIn("summaryCards.forEach((card) =>", html)
            self.assertIn("jumpButtons.forEach((button) =>", html)
            self.assertIn("const mainSummaryCards = summaryCards.filter", html)
            self.assertIn('activeFilter === nextFilter ? "" : nextFilter', html)
            self.assertIn("setFilter(\"\");", html)

    def test_html_meta_labels_sample_snapshots(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = self._snapshot(
                root,
                "[샘플] 작업 전",
                "GE1/0/1 UP",
                folder_label="sample_pre_work",
                stage_name="[샘플] 작업 전",
                stage_slug="sample_pre_work",
            )
            target = self._snapshot(
                root,
                "작업 전",
                "GE1/0/1 DOWN",
                folder_label="pre_work",
                stage_name="작업 전",
                stage_slug="pre_work",
            )
            summary = DiffEngine().compare(base, target)

            paths = ReportWriter().write_reports(summary)

            html = paths["html"].read_text(encoding="utf-8")
            self.assertIn(f"기준: 샘플: {base.name}", html)
            self.assertIn(f"비교: {target.name}", html)
            self.assertNotIn(f"비교: 샘플: {target.name}", html)

    def test_reports_redact_sensitive_values_but_keep_raw_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = self._snapshot(
                root,
                "base",
                "snmp-server community UltraSnmpSecret RO\npassword=OldSecret\nGE1/0/1 UP",
            )
            target = self._snapshot(
                root,
                "target",
                "snmp-server community UltraSnmpSecret RW\npassword=NewSecret\nGE1/0/1 DOWN",
            )
            summary = DiffEngine().compare(base, target)

            paths = ReportWriter().write_reports(summary)

            html = paths["html"].read_text(encoding="utf-8")
            manifest = json.loads(paths["json"].read_text(encoding="utf-8"))
            manifest_text = json.dumps(manifest, ensure_ascii=False)

            for secret in ("UltraSnmpSecret", "OldSecret", "NewSecret"):
                self.assertNotIn(secret, html)
                self.assertNotIn(secret, manifest_text)
            self.assertIn("***", html)
            self.assertIn("***", manifest_text)

            if "xlsx" in paths:
                from openpyxl import load_workbook

                workbook = load_workbook(paths["xlsx"])
                all_values = "\n".join(
                    str(cell)
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows(values_only=True)
                    for cell in row
                    if cell is not None
                )
                for secret in ("UltraSnmpSecret", "OldSecret", "NewSecret"):
                    self.assertNotIn(secret, all_values)

            base_raw = base / "raw" / "backbone4" / "interface_brief.txt"
            target_raw = target / "raw" / "backbone4" / "interface_brief.txt"
            self.assertIn("OldSecret", base_raw.read_text(encoding="utf-8"))
            self.assertIn("NewSecret", target_raw.read_text(encoding="utf-8"))

    def test_share_bundle_contains_redacted_reports_and_excludes_raw_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = self._snapshot(root, "base", "password=OldSecret\nGE1/0/1 UP")
            target = self._snapshot(root, "target", "password=NewSecret\nGE1/0/1 DOWN")
            summary = DiffEngine().compare(base, target)

            paths = ReportWriter().write_reports(summary)

            self.assertIn("share_zip", paths)
            self.assertTrue(paths["share_zip"].exists())
            with ZipFile(paths["share_zip"]) as archive:
                names = archive.namelist()
                self.assertIn("README_SHARED_REPORT.txt", names)
                self.assertIn("reports/diff_report.html", names)
                self.assertIn("reports/diff_manifest.json", names)
                self.assertTrue(
                    "reports/diff_summary.xlsx" in names or "reports/diff_summary.csv" in names
                )
                self.assertIn("docs/USER_GUIDE.md", names)
                self.assertIn("docs/COMMAND_GUIDE.md", names)
                self.assertIn("docs/COMMAND_GUIDE.html", names)
                self.assertIn("docs/images/settings-collection.png", names)
                self.assertIn("docs/images/compare-results.png", names)
                self.assertIn("docs/images/work-log.png", names)
                self.assertFalse(any("/raw/" in name or name.startswith("raw/") for name in names))
                self.assertFalse(any(name.endswith("devices.yaml") for name in names))
                self.assertFalse(any(name.endswith(".exe") for name in names))
                payload = "\n".join(
                    archive.read(name).decode("utf-8", errors="ignore")
                    for name in names
                    if name.endswith((".txt", ".html", ".json", ".md", ".csv"))
                )
            self.assertNotIn("OldSecret", payload)
            self.assertNotIn("NewSecret", payload)
            self.assertIn("***", payload)


if __name__ == "__main__":
    unittest.main()
